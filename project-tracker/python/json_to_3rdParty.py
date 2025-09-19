import json
import pandas as pd
import os
from openpyxl import load_workbook

def append_json_to_excel(json_path, excel_relative_path, sheet_name):
    # Load JSON data
    with open(json_path, 'r') as f:
        json_data = json.load(f)

    # Convert JSON data to DataFrame
    json_df = pd.DataFrame(json_data)

    # Define mapping of JSON fields to Excel columns
    field_mapping = {
        'Project Type': 'Project Type',
        'Engineering Number': 'Engineering Number',
        'WO Number': 'WO',
        'Town': 'Town',
        'Designer Name': 'Project Lead',
        'KH Billing No': 'KH Billing #',
        'LG Drawing': 'LG Drawing [Y or N]',
        '# of Design Pages': 'Sheets'
    }

    # Create a new DataFrame with mapped columns
    mapped_df = pd.DataFrame()
    for json_field, excel_column in field_mapping.items():
        mapped_df[excel_column] = json_df[json_field]

    # Load workbook and get the sheet
    excel_path = os.path.join(os.getcwd(), excel_relative_path)
    book = load_workbook(excel_path)
    writer = pd.ExcelWriter(excel_path, engine='openpyxl')
    writer.book = book

    # Get the last row in the existing sheet
    if sheet_name in book.sheetnames:
        sheet = book[sheet_name]
        startrow = sheet.max_row
    else:
        startrow = 0

    # Write the new data starting from the last row
    mapped_df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False, header=False)

    # Save the workbook
    writer.save()
    writer.close()


